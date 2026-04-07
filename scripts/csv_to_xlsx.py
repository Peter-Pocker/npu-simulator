#!/usr/bin/env python3
"""Convert NPU simulator CSV trace outputs to a single XLSX workbook."""

import argparse
import csv
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required.  Install with:\n  pip install openpyxl")
    sys.exit(1)

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
)

STATE_COLORS = {
    "IDLE":      PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "LOADING":   PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    "COMPUTING": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "WRITEBACK": PatternFill(start_color="E2BFEE", end_color="E2BFEE", fill_type="solid"),
    "DONE":      PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}


def _auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 60)


def _style_header(ws):
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
    ws.freeze_panes = "A2"


def _try_number(val: str):
    """Convert string to int or float when possible."""
    try:
        return int(val)
    except ValueError:
        pass
    try:
        return float(val)
    except ValueError:
        return val


def add_csv_sheet(wb: Workbook, sheet_name: str, csv_path: Path, color_states: bool = False):
    ws = wb.create_sheet(title=sheet_name)
    with open(csv_path, newline="") as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, start=1):
            for col_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx,
                               value=val if row_idx == 1 else _try_number(val))
                cell.border = THIN_BORDER

            if color_states and row_idx > 1:
                for col_idx, val in enumerate(row, start=1):
                    if val in STATE_COLORS:
                        ws.cell(row=row_idx, column=col_idx).fill = STATE_COLORS[val]

    _style_header(ws)
    _auto_width(ws)


def main():
    parser = argparse.ArgumentParser(description="Convert CSV trace files to XLSX")
    parser.add_argument("trace_dir", nargs="?", default="trace",
                        help="Directory containing CSV files (default: trace/)")
    parser.add_argument("-o", "--output", default=None,
                        help="Output XLSX path (default: <trace_dir>/trace_report.xlsx)")
    args = parser.parse_args()

    trace_dir = Path(args.trace_dir)
    if not trace_dir.is_dir():
        print(f"Error: directory '{trace_dir}' not found.")
        sys.exit(1)

    csv_files = sorted(trace_dir.glob("*.csv"))
    if not csv_files:
        print(f"Error: no CSV files found in '{trace_dir}'.")
        sys.exit(1)

    output = Path(args.output) if args.output else trace_dir / "trace_report.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    for csv_path in csv_files:
        sheet_name = csv_path.stem[:31]  # Excel sheet name limit
        color = "state_trace" in csv_path.stem
        add_csv_sheet(wb, sheet_name, csv_path, color_states=color)
        print(f"  + {sheet_name}  ({csv_path.name})")

    wb.save(str(output))
    print(f"\nSaved: {output}")


if __name__ == "__main__":
    main()
